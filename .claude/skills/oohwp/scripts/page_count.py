#!/usr/bin/env python3
"""
oohwp pagecount - HWP/HWPX/PDF 페이지 카운트 통합 스크립트.

지정 폴더 내 문서를 감지하여 변환 + 페이지 카운트 + xlsx 생성.

사용법:
    uv run --with pyhwpx,pypdf,openpyxl python page_count.py <폴더경로>

동작:
    1. 폴더 내 파일 타입 감지 (.hwp / .hwpx / .pdf)
    2. <폴더>/pagecount/ 하위 구조 생성
       - hwpx/  : HWP→HWPX 변환 결과 (원본이 .hwp인 경우)
       - pdf/   : PDF 변환 결과 (순번 prefix 포함: 01_파일명.pdf)
    3. pagecount/pagecount.xlsx 생성 (No, 파일명, 페이지수, 합계)
"""

import sys
import shutil
import argparse
from pathlib import Path

if sys.stdout.encoding != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8")


def p(msg):
    print(msg, flush=True)


def detect_source_type(src_dir: Path) -> str:
    """폴더 내 파일 타입 감지. hwp > hwpx > pdf 우선순위."""
    hwp = list(src_dir.glob("*.hwp"))
    hwpx = list(src_dir.glob("*.hwpx"))
    pdf = list(src_dir.glob("*.pdf"))

    if hwp:
        return "hwp"
    elif hwpx:
        return "hwpx"
    elif pdf:
        return "pdf"
    else:
        return "none"


def convert_hwp_to_hwpx(src_dir: Path, dst_dir: Path) -> list[Path]:
    """HWP → HWPX 변환. 변환된 hwpx 파일 목록 반환."""
    import pyhwpx

    hwp_files = sorted(src_dir.glob("*.hwp"))
    dst_dir.mkdir(parents=True, exist_ok=True)

    p(f"\n## HWP → HWPX 변환 ({len(hwp_files)}개)")
    results = []
    hwp = None
    try:
        hwp = pyhwpx.Hwp(visible=True)
        hwp.SetMessageBoxMode(0x10000)

        for i, hwp_file in enumerate(hwp_files, 1):
            hwpx_file = dst_dir / (hwp_file.stem + ".hwpx")
            try:
                result = hwp.Open(str(hwp_file.resolve()))
                if not result:
                    p(f"  [{i:2d}/{len(hwp_files)}] FAIL(열기): {hwp_file.name}")
                    continue
                hwp.SaveAs(str(hwpx_file.resolve()), format="HWPX")
                if hwpx_file.exists():
                    p(f"  [{i:2d}/{len(hwp_files)}] OK: {hwp_file.name}")
                    results.append(hwpx_file)
                else:
                    p(f"  [{i:2d}/{len(hwp_files)}] FAIL(저장): {hwp_file.name}")
            except Exception as e:
                p(f"  [{i:2d}/{len(hwp_files)}] FAIL: {hwp_file.name} - {e}")
    finally:
        if hwp:
            try:
                hwp.Quit()
            except Exception:
                pass

    p(f"  완료: {len(results)}/{len(hwp_files)}")
    return results


def convert_hwpx_to_pdf(src_dir: Path, dst_dir: Path) -> list[tuple[str, Path]]:
    """HWPX → PDF 변환. (원본명, pdf경로) 목록 반환. 순번 prefix 포함."""
    import pyhwpx

    hwpx_files = sorted(src_dir.glob("*.hwpx"))
    dst_dir.mkdir(parents=True, exist_ok=True)

    p(f"\n## HWPX → PDF 변환 ({len(hwpx_files)}개)")
    results = []
    hwp = None
    try:
        hwp = pyhwpx.Hwp(visible=True)
        hwp.SetMessageBoxMode(0x10000)

        for i, hwpx_file in enumerate(hwpx_files, 1):
            pdf_name = f"{i:02d}_{hwpx_file.stem}.pdf"
            pdf_file = dst_dir / pdf_name
            try:
                result = hwp.Open(str(hwpx_file.resolve()))
                if not result:
                    p(f"  [{i:2d}/{len(hwpx_files)}] FAIL(열기): {hwpx_file.name}")
                    continue
                hwp.SaveAs(str(pdf_file.resolve()), format="PDF")
                if pdf_file.exists():
                    p(f"  [{i:2d}/{len(hwpx_files)}] OK: {pdf_name}")
                    results.append((hwpx_file.stem, pdf_file))
                else:
                    p(f"  [{i:2d}/{len(hwpx_files)}] FAIL(저장): {hwpx_file.name}")
            except Exception as e:
                p(f"  [{i:2d}/{len(hwpx_files)}] FAIL: {hwpx_file.name} - {e}")
    finally:
        if hwp:
            try:
                hwp.Quit()
            except Exception:
                pass

    p(f"  완료: {len(results)}/{len(hwpx_files)}")
    return results


def copy_pdf_with_prefix(src_dir: Path, dst_dir: Path) -> list[tuple[str, Path]]:
    """PDF 파일을 순번 prefix 붙여 복사. (원본명, pdf경로) 목록 반환."""
    pdf_files = sorted(src_dir.glob("*.pdf"))
    dst_dir.mkdir(parents=True, exist_ok=True)

    p(f"\n## PDF 복사 ({len(pdf_files)}개)")
    results = []
    for i, pdf_file in enumerate(pdf_files, 1):
        pdf_name = f"{i:02d}_{pdf_file.name}"
        dst_file = dst_dir / pdf_name
        shutil.copy2(pdf_file, dst_file)
        p(f"  [{i:2d}/{len(pdf_files)}] {pdf_name}")
        results.append((pdf_file.stem, dst_file))

    return results


def count_pages_and_write_xlsx(
    pdf_entries: list[tuple[str, Path]], xlsx_path: Path
):
    """PDF 페이지 카운트 → xlsx 생성."""
    from pypdf import PdfReader
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side

    p(f"\n## 페이지 카운트")

    rows = []
    total_pages = 0
    for i, (orig_name, pdf_file) in enumerate(pdf_entries, 1):
        try:
            reader = PdfReader(str(pdf_file))
            pages = len(reader.pages)
            total_pages += pages
            rows.append((i, orig_name, pages))
            p(f"  [{i:2d}/{len(pdf_entries)}] {pages:3d}p  {orig_name}")
        except Exception as e:
            rows.append((i, orig_name, "오류"))
            p(f"  [{i:2d}/{len(pdf_entries)}] ERR  {orig_name} - {e}")

    # Excel 생성
    wb = Workbook()
    ws = wb.active
    ws.title = "페이지카운트"

    headers = ["No", "파일명", "페이지수"]
    header_font = Font(bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

    for row_idx, (no, name, pages) in enumerate(rows, 2):
        ws.cell(row=row_idx, column=1, value=no).border = thin_border
        ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=2, value=name).border = thin_border
        c = ws.cell(row=row_idx, column=3, value=pages)
        c.border = thin_border
        c.alignment = Alignment(horizontal="center")

    # 합계
    sum_row = len(rows) + 2
    ws.cell(row=sum_row, column=1, value="").border = thin_border
    c = ws.cell(row=sum_row, column=2, value="합계")
    c.font = Font(bold=True, size=11)
    c.border = thin_border
    c.alignment = Alignment(horizontal="center")
    c = ws.cell(row=sum_row, column=3, value=total_pages)
    c.font = Font(bold=True, size=11)
    c.border = thin_border
    c.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 65
    ws.column_dimensions["C"].width = 12

    wb.save(str(xlsx_path))

    p(f"\n## 결과: {len(rows)}개 파일, 총 {total_pages}페이지")
    p(f"   저장: {xlsx_path}")
    return total_pages


def main():
    parser = argparse.ArgumentParser(
        description="oohwp pagecount - 문서 페이지 카운트 (HWP/HWPX/PDF)"
    )
    parser.add_argument("folder", help="대상 폴더 경로 (.hwp/.hwpx/.pdf 자동 감지)")
    args = parser.parse_args()

    src_dir = Path(args.folder).resolve()
    if not src_dir.is_dir():
        p(f"[ERROR] 폴더 없음: {src_dir}")
        sys.exit(1)

    file_type = detect_source_type(src_dir)
    if file_type == "none":
        p(f"[ERROR] 지원 파일 없음 (.hwp/.hwpx/.pdf): {src_dir}")
        sys.exit(1)

    # pagecount 하위 구조 생성
    pc_dir = src_dir / "pagecount"
    pc_dir.mkdir(exist_ok=True)
    hwpx_dir = pc_dir / "hwpx"
    pdf_dir = pc_dir / "pdf"
    xlsx_path = pc_dir / "pagecount.xlsx"

    p(f"# oohwp pagecount")
    p(f"  소스: {src_dir}")
    p(f"  타입: {file_type}")
    p(f"  출력: {pc_dir}")

    pdf_entries: list[tuple[str, Path]] = []

    if file_type == "hwp":
        # HWP → HWPX → PDF
        convert_hwp_to_hwpx(src_dir, hwpx_dir)
        pdf_entries = convert_hwpx_to_pdf(hwpx_dir, pdf_dir)

    elif file_type == "hwpx":
        # HWPX → PDF
        pdf_entries = convert_hwpx_to_pdf(src_dir, pdf_dir)

    elif file_type == "pdf":
        # PDF 복사 (순번 prefix)
        pdf_entries = copy_pdf_with_prefix(src_dir, pdf_dir)

    if not pdf_entries:
        p("[ERROR] 변환된 PDF 없음")
        sys.exit(1)

    count_pages_and_write_xlsx(pdf_entries, xlsx_path)


if __name__ == "__main__":
    main()
